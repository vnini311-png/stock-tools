#!/usr/bin/env python3
"""修正 持股歷史股價_*.xlsx 的分頁命名（代號↔名稱錯置）。

背景：分頁名格式為「代號 名稱」，但有一批分頁的名稱與代號對不起來
（例：4979 被標成「矽格」，實際 4979 是華星光）。每日 append 是用開頭
4 碼代號比對，所以**資料本身正確**，只有標籤錯，人打開 xlsx 會誤讀。

權威名稱取自 twse_institutional_margin.csv 的「證券名稱」（TWSE/TPEx 官方）。
TWSE 的 `*` 註記（聖暉*、國巨*）視為同名，不改，以符合站上其他頁面的慣例。

用法:
    python3 tools/fix_price_sheet_names.py           # 只檢查，不寫入
    python3 tools/fix_price_sheet_names.py --apply   # 實際改名（原檔會先備份）
"""
import csv
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

DESKTOP = Path.home() / "Desktop/DESK/stock-tools"
CSV_PATH = DESKTOP / "claude code output/twse_institutional_margin.csv"
PREFIXES = ["持股歷史股價_30檔", "持股歷史股價_ADAM組合"]
SHEET_RE = re.compile(r"^(\d{4})\s+(.*)$")
INVALID_RE = re.compile(r"[\[\]:*?/\\]")


def official_names():
    names = {}
    with CSV_PATH.open(encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            names[r["股票代號"]] = r["證券名稱"].strip()
    return names


def latest(prefix):
    c = sorted(DESKTOP.glob(f"{prefix}-*.xlsx"))
    return c[-1] if c else None


def norm(s):
    """比對用：去掉 TWSE 的 * 註記與空白。"""
    return s.replace("*", "").strip()


def main():
    apply = "--apply" in sys.argv
    auth = official_names()
    total = 0
    for prefix in PREFIXES:
        path = latest(prefix)
        if not path:
            print(f"⚠ 找不到 {prefix}-*.xlsx")
            continue
        wb = load_workbook(path)
        fixes = []
        for sn in wb.sheetnames:
            m = SHEET_RE.match(sn)
            if not m:
                continue
            code, cur = m.group(1), m.group(2).strip()
            want = auth.get(code)
            if not want or norm(want) == norm(cur):
                continue
            # 保留現有的 * 慣例：官方有 * 而現有沒有時不算差異（已被 norm 濾掉）
            fixes.append((sn, INVALID_RE.sub("", f"{code} {want}")[:31]))
        print(f"\n{path.name}：{len(fixes)} 個分頁名需修正")
        for old, new in fixes:
            print(f"   {old:<18} → {new}")
        total += len(fixes)
        if fixes and apply:
            shutil.copy2(path, path.with_suffix(".xlsx.bak_names"))
            # 分兩階段，避免改名途中撞到彼此的舊名（本批含互換）
            for i, (old, _new) in enumerate(fixes):
                wb[old].title = f"__tmp{i}__"
            for i, (_old, new) in enumerate(fixes):
                wb[f"__tmp{i}__"].title = new
            wb.save(path)
            print(f"   ✓ 已寫入（備份 {path.name}.bak_names）")
        wb.close()

    if total and not apply:
        print(f"\n共 {total} 個分頁名不符。加 --apply 實際修正。")
    elif not total:
        print("\n✓ 所有分頁名皆正確")


if __name__ == "__main__":
    main()
