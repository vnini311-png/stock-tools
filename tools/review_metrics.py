#!/usr/bin/env python3
"""產生 data/review-metrics.json — 每日 Review 頁面自動帶入的量化數據。

資料來源（都是既有管線的產物，不另外爬）：
  1. ~/Desktop/DESK/stock-tools/持股歷史股價_{30檔,ADAM組合}-YYYYMMDD.xlsx
       → OHLCV → MA20/60/240、MACD(12,26,9)、KD(9,3,3)、RSI(14)、YTD/52週高低
  2. claude code output/twse_institutional_margin.csv
       → 三大法人買賣超、融資融券餘額與增減、融資使用率、券資比
  3. tools/chip_tool.html 的 dataPayload
       → 外資/投信/自營持股張數與比率、千張大戶占比
  4. TWSE TWT93U + TPEx margin/sbl（唯一需要連網的部分）
       → 借券賣出餘額 LS 與當日增減

用法:
    python3 tools/review_metrics.py            # 寫入 ../data/review-metrics.json
    python3 tools/review_metrics.py --push     # 順便 commit + push
    python3 tools/review_metrics.py --no-sbl   # 跳過借券（不連網）
"""
import csv
import json
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
DESKTOP = Path.home() / "Desktop/DESK/stock-tools"
CSV_PATH = DESKTOP / "claude code output/twse_institutional_margin.csv"
CHIP_HTML = REPO / "tools/chip_tool.html"
OUT_PATH = REPO / "data/review-metrics.json"

# 每日持股交易數據：兩個組合各一個檔，代號集合即為「目前持股」的範圍
PORTFOLIOS = [("SIVI", "持股歷史股價_30檔-*.xlsx"), ("ADAM", "持股歷史股價_ADAM組合-*.xlsx")]
MARKET_SHEET_RE = re.compile(r"大盤")
CODE_RE = re.compile(r"(\d{4})")


# ─────────────────── 指標計算 ───────────────────
def ema(series, span):
    k = 2 / (span + 1)
    out, prev = [], None
    for v in series:
        prev = v if prev is None else v * k + prev * (1 - k)
        out.append(prev)
    return out


def macd_series(closes, fast=12, slow=26, sig=9):
    if len(closes) < slow + sig:
        return None
    ef, es = ema(closes, fast), ema(closes, slow)
    dif = [a - b for a, b in zip(ef, es)]
    dea = ema(dif, sig)
    hist = [a - b for a, b in zip(dif, dea)]   # 與 macd.html 同慣例：OSC = DIF - MACD
    return dif, dea, hist


def kd_series(highs, lows, closes, n=9, a=3):
    """KD(9,3,3)，K/D 初始值 50。"""
    if len(closes) < n:
        return None, None
    ks, ds, k, d = [], [], 50.0, 50.0
    for i in range(len(closes)):
        if i < n - 1:
            ks.append(k)
            ds.append(d)
            continue
        hh = max(highs[i - n + 1:i + 1])
        ll = min(lows[i - n + 1:i + 1])
        rsv = 50.0 if hh == ll else (closes[i] - ll) / (hh - ll) * 100
        k = k + (rsv - k) / a
        d = d + (k - d) / a
        ks.append(k)
        ds.append(d)
    return ks, ds


def rsi_series(closes, n=14):
    """Wilder RSI(14)。"""
    if len(closes) < n + 1:
        return None
    gains = [max(closes[i] - closes[i - 1], 0) for i in range(1, len(closes))]
    losses = [max(closes[i - 1] - closes[i], 0) for i in range(1, len(closes))]
    ag = sum(gains[:n]) / n
    al = sum(losses[:n]) / n
    out = [None] * n
    out.append(100.0 if al == 0 else 100 - 100 / (1 + ag / al))
    for i in range(n, len(gains)):
        ag = (ag * (n - 1) + gains[i]) / n
        al = (al * (n - 1) + losses[i]) / n
        out.append(100.0 if al == 0 else 100 - 100 / (1 + ag / al))
    return out


def sma(values, n, i):
    if i + 1 < n:
        return None
    win = values[i - n + 1:i + 1]
    return sum(win) / n if all(v is not None for v in win) else None


def r(v, nd=2):
    return None if v is None else round(v, nd)


def pct(a, b):
    """a 相對 b 的變化率(%)。"""
    if a is None or not b:
        return None
    return round((a / b - 1) * 100, 2)


# ─────────────────── 讀價格 xlsx ───────────────────
def latest_price_files():
    files = []
    for tag, g in PORTFOLIOS:
        cand = sorted(DESKTOP.glob(g))
        if cand:
            files.append((tag, cand[-1]))
    if not files:
        sys.exit(f"找不到持股歷史股價 xlsx（{DESKTOP}）")
    return files


def read_sheet(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        d = row[0]
        d = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
        o, h, l, c = row[1], row[2], row[3], row[4]
        if c is None:
            continue                       # yfinance 偶發缺收盤 → 整列跳過
        v = row[5] if len(row) > 5 else None
        rows.append((d, o or c, h or c, l or c, c, v or 0))
    rows.sort(key=lambda x: x[0])
    return rows


def compute_tech(rows, ytd_base="2026-01-01"):
    if len(rows) < 30:
        return None
    dates = [x[0] for x in rows]
    highs = [x[2] for x in rows]
    lows = [x[3] for x in rows]
    closes = [x[4] for x in rows]
    vols = [x[5] for x in rows]
    i = len(rows) - 1

    t = {"date": dates[i]}
    t["close"] = r(closes[i])
    t["prev"] = r(closes[i - 1]) if i else None
    t["chg"] = r(closes[i] - closes[i - 1]) if i else None
    t["chgPct"] = pct(closes[i], closes[i - 1]) if i else None
    t["open"], t["high"], t["low"] = r(rows[i][1]), r(highs[i]), r(lows[i])
    t["volume"] = int(vols[i] or 0)
    v5 = sma(vols, 5, i)
    t["vol5"] = int(v5) if v5 else None
    t["volRatio"] = r(vols[i] / v5, 2) if v5 else None

    for n in (20, 60, 240):
        ma = sma(closes, n, i)
        t[f"ma{n}"] = r(ma)
        t[f"vsMa{n}"] = pct(closes[i], ma)

    m = macd_series(closes)
    if m:
        dif, dea, hist = m
        t["macd"], t["signal"] = r(dif[i], 4), r(dea[i], 4)
        t["hist"], t["histPrev"] = r(hist[i], 4), r(hist[i - 1], 4)
        t["trend"] = "多頭" if dif[i] > dea[i] else "空頭"
        t["position"] = "零軸上" if dif[i] > 0 else "零軸下"
        t["histDir"] = "擴大" if abs(hist[i]) > abs(hist[i - 1]) else "縮小"
        golden = dif[i] > dea[i] and dif[i - 1] <= dea[i - 1]
        dead = dif[i] < dea[i] and dif[i - 1] >= dea[i - 1]
        t["cross"] = "黃金交叉" if golden else ("死亡交叉" if dead else None)

    ks, ds = kd_series(highs, lows, closes)
    if ks:
        t["k"], t["d"] = r(ks[i]), r(ds[i])
        t["kPrev"], t["dPrev"] = r(ks[i - 1]), r(ds[i - 1])
        t["kdCross"] = ("黃金交叉" if ks[i] > ds[i] and ks[i - 1] <= ds[i - 1]
                        else "死亡交叉" if ks[i] < ds[i] and ks[i - 1] >= ds[i - 1] else None)
    rs = rsi_series(closes)
    if rs and rs[i] is not None:
        t["rsi"], t["rsiPrev"] = r(rs[i]), r(rs[i - 1])

    # YTD / 52 週高低
    ytd = [(d, hi, lo, c) for d, _o, hi, lo, c, _v in
           ((x[0], x[1], x[2], x[3], x[4], x[5]) for x in rows) if d >= ytd_base]
    if ytd:
        hd, hv = max(((d, hi) for d, hi, _lo, _c in ytd), key=lambda x: x[1])
        ld, lv = min(((d, lo) for d, _hi, lo, _c in ytd), key=lambda x: x[1])
        t["ytdHigh"], t["ytdHighDate"] = r(hv), hd
        t["ytdLow"], t["ytdLowDate"] = r(lv), ld
        t["fromHigh"] = pct(closes[i], hv)
        t["fromLow"] = pct(closes[i], lv)
        first_idx = next((j for j, d in enumerate(dates) if d >= ytd_base), None)
        if first_idx:
            t["ytdRet"] = pct(closes[i], closes[first_idx - 1])
    w52 = [x for x in rows if x[0] >= (datetime.strptime(dates[i], "%Y-%m-%d") - timedelta(days=365)).strftime("%Y-%m-%d")]
    if w52:
        t["high52"] = r(max(x[2] for x in w52))
        t["low52"] = r(min(x[3] for x in w52))
    return t


def load_prices():
    """回傳 {code: tech}, {code: 組合}, 大盤。代號集合 = 目前持股。"""
    out, port, market = {}, {}, None
    for tag, f in latest_price_files():
        wb = load_workbook(f, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            if MARKET_SHEET_RE.search(name):
                market = compute_tech(read_sheet(ws))
                if market:
                    market["name"] = "加權指數"
                continue
            m = CODE_RE.search(name)
            if not m:
                continue                       # 「📊 紅K量增分析」之類的說明分頁
            t = compute_tech(read_sheet(ws))
            if t:
                code = m.group(1)
                out[code] = t                  # 以代號為準：xlsx 分頁名的中文有錯置
                port[code] = tag
        wb.close()
    return out, port, market


# ─────────────────── 讀籌碼 CSV ───────────────────
def num(s):
    if s is None:
        return None
    s = str(s).replace(",", "").strip()
    if s in ("", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_chips_csv():
    """回傳 {code: [最新, 前一日]}，每筆是該股當日整列。"""
    by_code = {}
    with CSV_PATH.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            by_code.setdefault(row["股票代號"], []).append(row)
    out = {}
    for code, rows in by_code.items():
        rows.sort(key=lambda x: x["日期"])
        out[code] = rows[-2:]
    return out


def load_chip_payload():
    if not CHIP_HTML.exists():
        return {}
    h = CHIP_HTML.read_text(encoding="utf-8")
    m = re.search(r'<script[^>]*id="dataPayload"[^>]*>(.*?)</script>', h, re.S)
    return json.loads(m.group(1)) if m else {}


LOT = 1000.0

# ─────────────────── 借券賣出餘額（LS） ───────────────────
SBL_TWSE = "https://www.twse.com.tw/rwd/zh/marginTrading/TWT93U"
SBL_TPEX = "https://www.tpex.org.tw/www/zh-tw/margin/sbl"
UA = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"),
    "Accept": "application/json,text/plain,*/*",
}
# 兩張表的欄位都是 [代號, 名稱, 融券×6, 借券×6, 備註]，借券當日餘額在 index 12、前日在 8
SBL_PREV_I, SBL_BAL_I = 8, 12


def _get_json(url, params, referer, timeout=45):
    req = urllib.request.Request(f"{url}?{urllib.parse.urlencode(params)}",
                                 headers={**UA, "Referer": referer})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read())


def fetch_sbl(target_iso):
    """回傳 {code: (借券餘額張, 當日增減張)}。任何失敗都回空 dict，不擋主流程。"""
    out = {}
    ymd = target_iso.replace("-", "")
    slash = target_iso.replace("-", "/")

    def take(rows):
        for row in rows:
            cells = [str(c).strip() for c in row]
            code = cells[0]
            if not re.fullmatch(r"\d{4}", code):
                continue
            bal, prev = num(cells[SBL_BAL_I]), num(cells[SBL_PREV_I])
            if bal is None:
                continue
            out[code] = (round(bal / LOT), round((bal - (prev or 0)) / LOT))

    try:
        j = _get_json(SBL_TWSE, {"date": ymd, "selectType": "ALL", "response": "json"},
                      "https://www.twse.com.tw/")
        if j.get("stat") == "OK":
            take(j.get("data") or [])
        else:
            print(f"· 借券 TWSE {target_iso}: {j.get('stat')}")
    except Exception as e:
        print(f"⚠ 借券 TWSE 抓取失敗: {e}")

    time.sleep(3)
    try:
        j = _get_json(SBL_TPEX, {"date": slash, "response": "json"}, "https://www.tpex.org.tw/")
        for t in j.get("tables") or []:
            take(t.get("data") or [])
    except Exception as e:
        print(f"⚠ 借券 TPEx 抓取失敗: {e}")
    return out


def build_chips(code, csv_rows, chip, sbl=None):
    """對應筆記的 I 段（QFII/QDII/Trader/融資）與 S 段（融資UR/融券/P/C/LS）。"""
    cur = csv_rows[-1] if csv_rows else None
    if cur is None:
        return None, None, None
    snap = (chip or {}).get("snapshot") or {}
    issued = snap.get("issued_shares")
    issued_lots = issued / LOT if issued else None

    foreign_lots = (snap.get("foreign_shares") / LOT) if snap.get("foreign_shares") else None
    trust_lots = snap.get("trust_lots")
    dealer_lots = snap.get("dealer_lots")
    margin_lots = num(cur["融資融券_融資_今日餘額"])

    d_foreign = ((num(cur["外陸資買賣超股數(不含外資自營商)"]) or 0)
                 + (num(cur["外資自營商買賣超股數"]) or 0)) / LOT
    d_trust = (num(cur["投信買賣超股數"]) or 0) / LOT
    d_dealer = (num(cur["自營商買賣超股數"]) or 0) / LOT
    d_margin = (margin_lots or 0) - (num(cur["融資融券_融資_前日餘額"]) or 0)

    ratio = lambda lots: (lots / issued_lots) if (lots is not None and issued_lots) else None
    # I 段只放法人動向；融資/融券等信用交易一律歸 S 段
    I = {
        "labels": ["QFII", "QDII", "Trader"],
        "value": [r(foreign_lots, 0), r(trust_lots, 0), r(dealer_lots, 0)],
        "change": [r(d_foreign, 0), r(d_trust, 0), r(d_dealer, 0)],
        "ratio": [r((num(cur["外資及陸資持股比率(%)"]) or 0) / 100, 4),
                  r(ratio(trust_lots), 4), r(ratio(dealer_lots), 4)],
    }

    quota = num(cur["融資融券_融資_次一營業日限額"])
    short_lots = num(cur["融資融券_融券_今日餘額"])
    short_prev = num(cur["融資融券_融券_前日餘額"])
    ur = (margin_lots / quota) if (margin_lots is not None and quota) else None
    pc = (short_lots / margin_lots * 100) if (short_lots is not None and margin_lots) else None

    ur_prev = pc_prev = None
    if len(csv_rows) > 1:
        p = csv_rows[0]
        mp, qp, sp = num(p["融資融券_融資_今日餘額"]), num(p["融資融券_融資_次一營業日限額"]), num(p["融資融券_融券_今日餘額"])
        ur_prev = (mp / qp) if (mp is not None and qp) else None
        pc_prev = (sp / mp * 100) if (sp is not None and mp) else None

    ls_bal, ls_chg = sbl if sbl else (None, None)
    # 融資使用率 UR = 融資餘額 / 次一營業日限額（限額為發行股數的 25%），
    # 與「融資占發行股數比率」是同一件事的兩種刻度（UR = 占股本% × 4），這裡放前者。
    S = {
        "labels": ["融資", "融資UR", "融券", "P/C", "LS"],
        "value": [r(margin_lots, 0), r(ur, 4), r(short_lots, 0), r(pc, 2), ls_bal],
        "change": [r(d_margin, 0),
                   r(ur - ur_prev, 4) if (ur is not None and ur_prev is not None) else None,
                   r(short_lots - short_prev, 0) if (short_lots is not None and short_prev is not None) else None,
                   r(pc - pc_prev, 2) if (pc is not None and pc_prev is not None) else None,
                   ls_chg],
    }

    big = build_big(chip)
    return I, S, big


def build_big(chip):
    """千張大戶：TDCC 每週五一筆。回傳最新值、週增減、連續同向週數與近 8 週序列。"""
    ser = (chip or {}).get("series") or {}
    pcts, dates = ser.get("big_pct"), ser.get("dates")
    if not pcts or not dates:
        return None
    weeks = [(d, v) for d, v in zip(dates, pcts) if v is not None]
    if not weeks:
        return None
    last_d, last_v = weeks[-1]
    out = {"pct": r(last_v), "date": last_d, "chg": None, "streak": 0, "dir": None,
           "weeks": [[d, r(v)] for d, v in weeks[-8:]]}
    if len(weeks) < 2:
        return out
    diffs = [round(weeks[i][1] - weeks[i - 1][1], 4) for i in range(1, len(weeks))]
    out["chg"] = r(diffs[-1])
    if diffs[-1] != 0:
        sign = 1 if diffs[-1] > 0 else -1
        n = 0
        for dv in reversed(diffs):
            if (dv > 0 and sign > 0) or (dv < 0 and sign < 0):
                n += 1
            else:
                break
        out["streak"], out["dir"] = n, ("增" if sign > 0 else "減")
    return out


# ─────────────────── 主流程 ───────────────────
def main():
    prices, port, market = load_prices()
    csv_chips = load_chips_csv()
    payload = load_chip_payload()

    chip_date = max((rows[-1]["日期"] for rows in csv_chips.values() if rows), default=None)
    sbl = {} if "--no-sbl" in sys.argv else fetch_sbl(chip_date) if chip_date else {}

    names = {}
    for code, rows in csv_chips.items():
        names[code] = (rows[-1].get("證券名稱") or "").strip()
    for code, v in payload.items():
        if v.get("name"):
            names[code] = v["name"]

    # 範圍限縮在「每日持股交易數據」涵蓋的個股 = 我的持股 + Adam 持股
    stocks = {}
    for code in sorted(prices):
        tech = prices[code]
        I, S, big = build_chips(code, csv_chips.get(code, []), payload.get(code), sbl.get(code))
        stocks[code] = {
            "code": code,
            "name": names.get(code) or code,
            "portfolio": port.get(code),
            "tech": tech,
            "chips": I,
            "credit": S,
            "big": big,
            "chipDate": csv_chips.get(code, [{}])[-1].get("日期"),
        }

    counts = {}
    for v in stocks.values():
        counts[v["portfolio"]] = counts.get(v["portfolio"], 0) + 1
    doc = {
        "portfolios": counts,
        "generated": datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S+08:00"),
        "priceDate": market["date"] if market else None,
        "chipDate": max((s["chipDate"] for s in stocks.values() if s["chipDate"]), default=None),
        "market": market,
        "stocks": stocks,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(doc, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    miss = [c for c, v in stocks.items() if not v["chips"]]
    print(f"✓ {OUT_PATH.relative_to(REPO)}  {len(stocks)} 檔 "
          f"({', '.join(f'{k} {v}' for k, v in sorted(counts.items()))}) "
          f"(股價 {doc['priceDate']} / 籌碼 {doc['chipDate']})  {OUT_PATH.stat().st_size // 1024} KB")
    if miss:
        print(f"⚠ 無籌碼資料: {', '.join(miss)}（CSV 缺該代號）")
    no_ls = [c for c, v in stocks.items() if (v["credit"] or {}).get("value", [None] * 5)[4] is None]
    if no_ls:
        print(f"· 無借券餘額: {', '.join(no_ls)}")

    if "--push" in sys.argv:
        subprocess.run(["git", "-C", str(REPO), "add", str(OUT_PATH.relative_to(REPO))], check=True)
        msg = f"Update review metrics through {doc['priceDate']} close"
        if subprocess.run(["git", "-C", str(REPO), "diff", "--cached", "--quiet"]).returncode:
            subprocess.run(["git", "-C", str(REPO), "commit", "-m", msg], check=True)
            subprocess.run(["git", "-C", str(REPO), "pull", "--rebase"], check=True)
            subprocess.run(["git", "-C", str(REPO), "push"], check=True)
            print("✓ pushed")
        else:
            print("· 無變更，不 commit")


if __name__ == "__main__":
    main()
